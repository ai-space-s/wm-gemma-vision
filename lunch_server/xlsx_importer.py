from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


class XlsxExtractionError(RuntimeError):
    pass


MEAL_ROW_TO_KEY = {"조식": "breakfast", "중식": "lunch", "석식": "dinner"}
MEAL_KEYS = ("breakfast", "lunch", "dinner")
SOUP_KEYWORDS = ("국", "탕", "찌개", "스프", "우동", "쌀국수", "수제비", "곰탕", "냉면")
DRINK_KEYWORDS = ("주스", "쥬스", "우유", "두유", "요구르트", "요거트", "라떼", "식혜", "브리즈")
DESSERT_KEYWORDS = (
    "과일",
    "바나나",
    "딸기",
    "오렌지",
    "파인애플",
    "키위",
    "사과",
    "케이크",
    "푸딩",
    "젤리",
    "도넛",
    "파이",
)


def parse_xlsx_meal_menus(
    path: Path,
    *,
    source_filename: str | None = None,
) -> list[dict[str, Any]]:
    if path.suffix.lower() != ".xlsx":
        raise XlsxExtractionError("xlsx 파일만 가져올 수 있습니다.")

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise XlsxExtractionError(f"xlsx 파일을 열지 못했습니다: {exc}") from exc

    records_by_date: dict[str, dict[str, Any]] = {}
    display_filename = source_filename or path.name
    for worksheet in workbook.worksheets:
        year, month = _find_year_month(worksheet, path)
        _parse_sheet(worksheet, year, month, display_filename, records_by_date)

    if not records_by_date:
        raise XlsxExtractionError("xlsx에서 식단을 찾지 못했습니다.")
    return [records_by_date[key] for key in sorted(records_by_date)]


def _parse_sheet(
    worksheet: Worksheet,
    year: int,
    month: int,
    source_filename: str,
    records_by_date: dict[str, dict[str, Any]],
) -> None:
    for row in range(2, worksheet.max_row + 1):
        label = _meal_label_for_row(worksheet, row)
        if label is None:
            continue

        date_row = _find_date_row_above(worksheet, row)
        if date_row is None:
            continue

        meal_key = MEAL_ROW_TO_KEY[label]
        for column in range(1, worksheet.max_column + 1):
            day = _parse_day(_cell_value(worksheet, date_row, column))
            if day is None:
                continue

            try:
                target_date = date(year, month, day)
            except ValueError:
                continue

            key = target_date.isoformat()
            record = records_by_date.setdefault(
                key,
                {
                    "date": key,
                    "meals": {meal: _empty_meal() for meal in MEAL_KEYS},
                    "source": f"xlsx:{source_filename}:{worksheet.title}",
                },
            )
            raw_text = _clean_text(_cell_value(worksheet, row, column))
            if raw_text:
                record["meals"][meal_key] = {
                    "hasMeal": True,
                    "reason": "",
                    "reasonCode": "",
                    "menu": _menu_from_text(raw_text),
                }
            else:
                record["meals"][meal_key] = {
                    "hasMeal": False,
                    "reason": f"식단표에 {label}이 없습니다.",
                    "reasonCode": "no_meal_in_sheet",
                    "menu": _empty_menu(),
                }


def _meal_label_for_row(worksheet: Worksheet, row: int) -> str | None:
    for column in range(1, worksheet.max_column + 1):
        value = _clean_text(_cell_value(worksheet, row, column))
        if value in MEAL_ROW_TO_KEY:
            return value
    return None


def _find_date_row_above(worksheet: Worksheet, meal_row: int) -> int | None:
    candidates: list[tuple[int, int]] = []
    for row in range(max(1, meal_row - 4), meal_row):
        day_count = sum(
            1
            for column in range(1, worksheet.max_column + 1)
            if _parse_day(_cell_value(worksheet, row, column)) is not None
        )
        if day_count:
            candidates.append((day_count, row))
    if not candidates:
        return None
    return max(candidates)[1]


def _find_year_month(worksheet: Worksheet, path: Path) -> tuple[int, int]:
    for row in worksheet.iter_rows(values_only=True):
        for value in row:
            text = _clean_text(value)
            match = re.search(r"(\d{4})\s*년\s*(\d{1,2})\s*월", text)
            if match:
                return int(match.group(1)), int(match.group(2))

    match = re.search(r"(\d{4})년\s*(\d{1,2})월", path.name)
    if match:
        return int(match.group(1)), int(match.group(2))

    raise XlsxExtractionError("조회년월을 찾지 못했습니다.")


def _cell_value(worksheet: Worksheet, row: int, column: int) -> Any:
    cell = worksheet.cell(row, column)
    if not isinstance(cell, MergedCell):
        return cell.value

    coordinate = cell.coordinate
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _parse_day(value: Any) -> int | None:
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float) and value.is_integer():
        day = int(value)
        return day if 1 <= day <= 31 else None

    text = _clean_text(value)
    if not text:
        return None
    match = re.fullmatch(r"(\d{1,2})", text)
    if not match:
        return None
    day = int(match.group(1))
    return day if 1 <= day <= 31 else None


def _menu_from_text(raw_text: str) -> dict[str, Any]:
    items = _split_menu_items(raw_text)
    soup = _first_matching(items, SOUP_KEYWORDS)
    main = next((item for item in items if item != soup), "")
    tail_candidates = [item for item in items if item not in {main, soup}]
    drink = _last_matching(tail_candidates, DRINK_KEYWORDS)
    dessert = _last_matching(
        [item for item in tail_candidates if item != drink],
        DESSERT_KEYWORDS,
    )
    excluded = {value for value in (main, soup, drink, dessert) if value}

    return {
        "main": main,
        "soup": soup,
        "sideDishes": [item for item in items if item not in excluded],
        "dessert": dessert,
        "drink": drink,
        "items": items,
        "notes": "",
        "rawText": raw_text,
    }


def _split_menu_items(raw_text: str) -> list[str]:
    normalized = raw_text.replace("\r", "\n").replace("\n", "/")
    parts = [_clean_text(part) for part in normalized.split("/")]
    items: list[str] = []
    for part in parts:
        if not part:
            continue
        if part.startswith("*") or "에너지" in part:
            break
        if re.fullmatch(r"[\d,.]+", part):
            continue
        items.append(part)
    return items


def _first_matching(items: list[str], keywords: tuple[str, ...]) -> str:
    return next((item for item in items if any(keyword in item for keyword in keywords)), "")


def _last_matching(items: list[str], keywords: tuple[str, ...]) -> str:
    return next(
        (item for item in reversed(items) if any(keyword in item for keyword in keywords)),
        "",
    )


def _empty_meal() -> dict[str, Any]:
    return {
        "hasMeal": False,
        "reason": "식단표에 식사가 없습니다.",
        "reasonCode": "no_meal_in_sheet",
        "menu": _empty_menu(),
    }


def _empty_menu() -> dict[str, Any]:
    return {
        "main": "",
        "soup": "",
        "sideDishes": [],
        "dessert": "",
        "drink": "",
        "items": [],
        "notes": "",
        "rawText": "",
    }


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
